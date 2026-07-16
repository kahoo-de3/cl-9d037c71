import json, sys, os, base64, hashlib
import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

# Usage: python generate.py <Book.xlsx> <PIN>  -> writes docs/index.html
# データはPIN由来の鍵でAES-GCM暗号化して埋め込む（PIN自体はどこにも保存されない）
xlsx = sys.argv[1]
pin = sys.argv[2]
PBKDF2_ITER = 310000
# 外字（私用領域）→通常漢字の置換表。新しい外字が出たら警告を見てここに追加する
PUA_MAP = {chr(0xE682): chr(0x7950)}  # 祐

# 施設名→読み（ひらがな）。五十音順ソート用。新しい施設が出たら警告を見てここに追加する
YOMI = {
    '我孫子東邦病院　医療': 'あびことうほうびょういん',
    '梅郷整形外科クリニック　医療': 'うめさとせいけいげかくりにっく',
    'おおすか整形外科　医療': 'おおすかせいけいげか',
    'おおたかの森病院　医療': 'おおたかのもりびょういん',
    '柏整形外科クリニック　医療': 'かしわせいけいげかくりにっく',
    '柏たなか病院　医療': 'かしわたなかびょういん',
    '柏の葉整形外科リハビリＣＬ': 'かしわのはせいけいげかりはびりくりにっく',
    '柏病院　市立': 'かしわびょういん',
    '鎌ケ谷総合病院　医療': 'かまがやそうごうびょういん',
    '川間春日町整形外科小児科ＣＬ': 'かわまかすがちょうせいけいげかしょうにかくりにっく',
    '北習志野整形外科クリニック': 'きたならしのせいけいげかくりにっく',
    '北習志野花輪病院　医療': 'きたならしのはなわびょういん',
    'キッコーマン総合病院': 'きっこーまんそうごうびょういん',
    '国立がん研究センター東病院': 'こくりつがんけんきゅうせんたーひがしびょういん',
    '五香整形外科内科　医療': 'ごこうせいけいげかないか',
    '五香病院　医療': 'ごこうびょういん',
    '逆井記念医院': 'さかさいきねんいいん',
    'さかの整形外科クリニック': 'さかのせいけいげかくりにっく',
    'さなだクリニック　医療': 'さなだくりにっく',
    '新東京クリニック　医療': 'しんとうきょうくりにっく',
    '新松戸むらた整形外科　医療': 'しんまつどむらたせいけいげか',
    '新八柱整形外科内科　医療': 'しんやはしらせいけいげかないか',
    'たけぐち整形外科　医療': 'たけぐちせいけいげか',
    'ためがい整形外科クリニック': 'ためがいせいけいげかくりにっく',
    'ちよだ整形外科　医療': 'ちよだせいけいげか',
    '東京慈恵会医科大学附属柏病院': 'とうきょうじけいかいいかだいがくふぞくかしわびょういん',
    '東邦鎌谷病院　医療': 'とうほうかまがやびょういん',
    'なかじま整形外科クリニック': 'なかじませいけいげかくりにっく',
    '流山中央病院　医療': 'ながれやまちゅうおうびょういん',
    '名戸ケ谷病院　医療': 'などがやびょういん',
    'はた整形外科　医療': 'はたせいけいげか',
    '原木中山駅前整形外科・リハビリ': 'ばらきなかやまえきまえせいけいげかりはびり',
    '船橋市立医療センター': 'ふなばししりついりょうせんたー',
    '船橋整形外科西船クリニック': 'ふなばしせいけいげかにしふなくりにっく',
    '船橋整形外科病院　医療': 'ふなばしせいけいげかびょういん',
    '船橋二和病院　医療': 'ふなばしふたわびょういん',
    '平和台病院　医療': 'へいわだいびょういん',
    'ますお整形外科・内科': 'ますおせいけいげかないか',
    '松戸市立総合医療センター': 'まつどしりつそうごういりょうせんたー',
    '松戸整形外科クリニック　医療': 'まつどせいけいげかくりにっく',
    '松戸整形外科病院　医療': 'まつどせいけいげかびょういん',
    '松戸常盤平いいだ整形外科　医療': 'まつどときわだいらいいだせいけいげか',
    '稔台整形外科クリニック': 'みのりだいせいけいげかくりにっく',
    'もり整形外科クリニック　医療': 'もりせいけいげかくりにっく',
}
wb = openpyxl.load_workbook(xlsx)
ws = wb.active
cells = {}
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for c in row:
        if c.value is not None:
            cells[c.coordinate] = str(c.value)

def fix_pua(text, coord):
    out = []
    for ch in text:
        if ch in PUA_MAP:
            out.append(PUA_MAP[ch])
        elif 0xE000 <= ord(ch) <= 0xF8FF:
            print(f'警告: {coord} に未対応の外字 U+{ord(ch):04X} があります。'
                  f'Web上で表示できないため PUA_MAP に置換先を追加してください。')
            out.append(ch)
        else:
            out.append(ch)
    return ''.join(out)

rows = []
for r in range(5, ws.max_row + 1):
    b = cells.get(f'B{r}')
    if not b:
        continue
    rows.append({
        'r': r,
        'f': fix_pua(b.strip(), f'B{r}'),
        'n': fix_pua((cells.get(f'C{r}') or '').strip(), f'C{r}'),
        'c': f'D{r}' in cells,
        'o': f'E{r}' in cells,
    })

# 施設名の読み（五十音順）でソート。同一施設は元のExcel行順を維持
_SMALL = {'ぁ':'あ','ぃ':'い','ぅ':'う','ぇ':'え','ぉ':'お','っ':'つ','ゃ':'や','ゅ':'ゆ','ょ':'よ','ゎ':'わ'}
def yomi_key(f):
    y = YOMI.get(f)
    if y is None:
        print(f'警告: 施設「{f}」の読みが YOMI にありません。読みを追加してください（暫定で末尾に並べます）。')
        return 'ん' * 20 + f
    return ''.join(_SMALL.get(ch, ch) for ch in y if ch != 'ー')
rows.sort(key=lambda x: (yomi_key(x['f']), x['r']))

data_js = json.dumps(rows, ensure_ascii=False, separators=(',', ':'))

salt = os.urandom(16)
iv = os.urandom(12)
key = hashlib.pbkdf2_hmac('sha256', pin.encode('utf-8'), salt, PBKDF2_ITER, dklen=32)
ct = AESGCM(key).encrypt(iv, data_js.encode('utf-8'), None)
enc_js = json.dumps({
    'salt': base64.b64encode(salt).decode(),
    'iv': base64.b64encode(iv).decode(),
    'ct': base64.b64encode(ct).decode(),
    'iter': PBKDF2_ITER,
    'len': len(pin),
}, separators=(',', ':'))

print('rows:', len(rows))
print('C-only:', sum(1 for x in rows if x['c'] and not x['o']))
print('O-only:', sum(1 for x in rows if x['o'] and not x['c']))
print('both:', sum(1 for x in rows if x['c'] and x['o']))

html = """<meta charset="utf-8">
<meta name="robots" content="noindex, nofollow">
<title>C/O チェックリスト</title>
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1">
<style>
/* ユーザー指定により常に白背景（ダークモードでも固定） */
:root{
  --bg:#FFFFFF; --card:#FFFFFF; --ink:#1D1D1F; --sub:#6D6D70;
  --line:#E2E2DE; --accent:#0B7285; --accent-ink:#FFFFFF;
  --partial-bg:#E9E9E6; --partial-ink:#9A9A9C;
  --done-bg:#66666A; --done-ink:#D9D9DC;
  --btn-bg:#FFFFFF; --btn-line:#C9C9C4;
  color-scheme: light;
}
*{box-sizing:border-box; -webkit-tap-highlight-color:transparent;}
[hidden]{display:none !important;}
html,body{margin:0;}
body{
  background:var(--bg); color:var(--ink);
  font-family:"Hiragino Sans","Hiragino Kaku Gothic ProN","Yu Gothic UI","Yu Gothic",Meiryo,sans-serif;
  font-size:15px; line-height:1.45;
}
header{
  position:sticky; top:0; z-index:10;
  background:var(--bg); border-bottom:1px solid var(--line);
  padding:10px 14px 8px;
}
.hrow{display:flex; align-items:center; gap:8px; flex-wrap:wrap;}
h1{font-size:15px; font-weight:700; margin:0 auto 0 0; letter-spacing:.02em;}
select{
  font:inherit; font-weight:600; color:var(--ink);
  background:var(--card); border:1px solid var(--btn-line); border-radius:8px;
  padding:7px 8px; min-height:38px;
}
.hrow2{display:flex; align-items:center; gap:10px; margin-top:8px;}
.progress{flex:1; height:8px; border-radius:4px; background:var(--partial-bg); overflow:hidden;}
.progress>div{height:100%; width:0%; background:var(--accent); border-radius:4px; transition:width .2s;}
.count{font-variant-numeric:tabular-nums; font-weight:700; font-size:13px; white-space:nowrap;}
.count small{font-weight:500; color:var(--sub);}
.hrow3{display:flex; gap:8px; margin-top:8px;}
.chip{
  font:inherit; font-size:13px; font-weight:600;
  border:1px solid var(--btn-line); background:var(--card); color:var(--ink);
  border-radius:999px; padding:6px 14px; min-height:34px;
}
.chip.on{background:var(--accent); border-color:var(--accent); color:var(--accent-ink);}
.chip.reset{margin-left:auto; color:var(--sub);}
main{padding:8px 10px 60px; max-width:640px; margin:0 auto;}
.row{
  display:flex; align-items:center; gap:10px;
  background:var(--card); border:1px solid var(--line); border-radius:12px;
  padding:9px 10px 9px 12px; margin-bottom:6px;
}
.idx{font-size:11px; color:var(--sub); font-variant-numeric:tabular-nums; width:2em; flex:none; text-align:right;}
.txt{flex:1; min-width:0;}
.fac{font-weight:600; font-size:14px; overflow-wrap:anywhere;}
.nm{font-size:13px; color:var(--sub);}
.btns{display:flex; gap:8px; flex:none;}
button.co{
  font:inherit; font-weight:800; font-size:16px;
  width:46px; height:46px; border-radius:10px;
  border:1.5px solid var(--btn-line); background:var(--btn-bg); color:var(--ink);
  display:flex; align-items:center; justify-content:center;
}
button.co:active{transform:scale(.94);}
button.co.on{background:var(--accent); border-color:var(--accent); color:var(--accent-ink);}
.spacer{width:46px; height:46px; flex:none;}
/* partial tap -> light gray */
.row.partial{background:var(--partial-bg);}
.row.partial .fac, .row.partial .nm, .row.partial .idx{color:var(--partial-ink);}
/* all tapped -> dark gray */
.row.done{background:var(--done-bg); border-color:var(--done-bg);}
.row.done .fac, .row.done .nm, .row.done .idx{color:var(--done-ink);}
.row.done .fac{text-decoration:line-through; text-decoration-thickness:1px;}
.row.done button.co.on{background:transparent; border-color:var(--done-ink); color:var(--done-ink);}
.empty{color:var(--sub); text-align:center; padding:40px 0; font-size:14px;}
.banner{
  margin-top:8px; padding:8px 10px; border-radius:8px;
  background:#B4232310; border:1px solid #B42323; color:#B42323;
  font-size:12px; line-height:1.5;
}
.lock{
  display:flex; justify-content:center;
  padding:14vh 20px 40px;
}
.lockbox{display:flex; flex-direction:column; gap:12px; width:min(280px, 80vw); text-align:center;}
.lockver{font-size:11px; color:var(--sub); margin-top:16px;}
.locktitle{font-weight:700; font-size:16px;}
.lockbox input{
  font:inherit; font-size:22px; text-align:center; letter-spacing:.3em;
  padding:10px; border:1.5px solid var(--btn-line); border-radius:10px;
  background:var(--card); color:var(--ink); width:100%;
}
.lockbox .chip{font-size:16px; padding:12px 14px; min-height:48px;}
.lockbox .chip:disabled{opacity:.6;}
.lockerr{color:#B42323; font-size:13px;}
@media (prefers-reduced-motion: reduce){ .progress>div{transition:none;} button.co:active{transform:none;} }
</style>

<div class="lock" id="lock" hidden>
  <form class="lockbox" id="lockform">
    <div class="locktitle">🔒 PINを入力してください</div>
    <input id="pin" type="password" inputmode="numeric" enterkeyhint="go" autocomplete="off" placeholder="暗証番号">
    <button class="chip" id="unlock" type="submit">開く</button>
    <div class="lockerr" id="lockerr" hidden>PINが違います</div>
    <div class="lockver">v3 ・ 最後の桁まで入力すると自動で開きます</div>
  </form>
</div>

<header>
  <div class="hrow">
    <h1>C/O チェック</h1>
    <select id="year" aria-label="年"></select>
    <select id="month" aria-label="月"></select>
  </div>
  <div class="hrow2">
    <div class="progress"><div id="bar"></div></div>
    <div class="count"><span id="done">0</span><small> / <span id="total">–</span> 完了</small></div>
  </div>
  <div class="hrow3">
    <button class="chip" id="filter">未完了のみ表示</button>
    <button class="chip reset" id="reset">リセット</button>
  </div>
  <div class="banner" id="banner" hidden>⚠ この環境ではチェック内容の自動保存が使えないため、閉じると消えます。</div>
</header>
<main id="list"></main>
<div class="empty" id="empty" hidden>この月は全件チェック済みです 🎉</div>

<script>
const ENC = __ENC__;
let DATA = [];
let TOTAL = 0;

const yearSel = document.getElementById('year');
const monthSel = document.getElementById('month');
for (let y = 2026; y <= 2030; y++) yearSel.add(new Option(y + '年', y));
for (let m = 1; m <= 12; m++) monthSel.add(new Option(m + '月', m));
yearSel.value = '2026'; monthSel.value = '7';

let state = {};
let filterOn = false;

// storage layer: localStorage if usable, otherwise in-memory (with warning banner)
let storageOK = true;
try {
  localStorage.setItem('co-check-test', '1');
  if (localStorage.getItem('co-check-test') !== '1') throw 0;
  localStorage.removeItem('co-check-test');
} catch(e){ storageOK = false; }

const memStore = {};
function storeGet(key){
  if (storageOK){ try { return localStorage.getItem(key); } catch(e){} }
  return Object.prototype.hasOwnProperty.call(memStore, key) ? memStore[key] : null;
}
function storeSet(key, val){
  if (storageOK){ try { localStorage.setItem(key, val); return; } catch(e){ storageOK = false; showBanner(); } }
  memStore[key] = val;
}
function showBanner(){ document.getElementById('banner').hidden = false; }
if (!storageOK) showBanner();

function storageKey(){ return 'co-check-' + yearSel.value + '-' + monthSel.value; }
function load(){
  try { state = JSON.parse(storeGet(storageKey())) || {}; }
  catch(e){ state = {}; }
}
function save(){
  storeSet(storageKey(), JSON.stringify(state));
}

function rowStatus(item){
  const s = state[item.r] || {};
  const need = (item.c ? 1 : 0) + (item.o ? 1 : 0);
  const got = (item.c && s.c ? 1 : 0) + (item.o && s.o ? 1 : 0);
  if (got === 0) return 'none';
  return got >= need ? 'done' : 'partial';
}

const list = document.getElementById('list');
const rowEls = {};

function build(){
  list.textContent = '';
  for (const item of DATA){
    const row = document.createElement('div');
    row.className = 'row';
    const idx = document.createElement('span');
    idx.className = 'idx'; idx.textContent = item.r;
    const txt = document.createElement('div');
    txt.className = 'txt';
    const fac = document.createElement('div');
    fac.className = 'fac'; fac.textContent = item.f;
    const nm = document.createElement('div');
    nm.className = 'nm'; nm.textContent = item.n;
    txt.append(fac, nm);
    const btns = document.createElement('div');
    btns.className = 'btns';
    for (const k of ['c','o']){
      if (item[k]){
        const b = document.createElement('button');
        b.className = 'co'; b.dataset.k = k;
        b.textContent = k.toUpperCase();
        b.setAttribute('aria-pressed', 'false');
        b.addEventListener('click', () => toggle(item, k));
        btns.appendChild(b);
      } else {
        const sp = document.createElement('span');
        sp.className = 'spacer';
        btns.appendChild(sp);
      }
    }
    row.append(idx, txt, btns);
    list.appendChild(row);
    rowEls[item.r] = row;
  }
}

function toggle(item, k){
  const s = state[item.r] || (state[item.r] = {});
  s[k] = !s[k];
  save();
  render();
}

function render(){
  let doneCount = 0;
  for (const item of DATA){
    const row = rowEls[item.r];
    const st = rowStatus(item);
    row.classList.toggle('partial', st === 'partial');
    row.classList.toggle('done', st === 'done');
    if (st === 'done') doneCount++;
    const s = state[item.r] || {};
    for (const b of row.querySelectorAll('button.co')){
      const on = !!s[b.dataset.k];
      b.classList.toggle('on', on);
      b.setAttribute('aria-pressed', String(on));
    }
    row.hidden = filterOn && st === 'done';
  }
  document.getElementById('done').textContent = doneCount;
  document.getElementById('bar').style.width = (TOTAL ? doneCount / TOTAL * 100 : 0) + '%';
  document.getElementById('empty').hidden = !(filterOn && doneCount === TOTAL);
}

yearSel.addEventListener('change', () => { load(); render(); });
monthSel.addEventListener('change', () => { load(); render(); });

document.getElementById('filter').addEventListener('click', (e) => {
  filterOn = !filterOn;
  e.currentTarget.classList.toggle('on', filterOn);
  render();
});
document.getElementById('reset').addEventListener('click', () => {
  if (confirm(yearSel.value + '\\u5e74' + monthSel.value + '\\u6708\\u306e\\u30c1\\u30a7\\u30c3\\u30af\\u3092\\u3059\\u3079\\u3066\\u30ea\\u30bb\\u30c3\\u30c8\\u3057\\u307e\\u3059\\u304b\\uff1f')){
    state = {}; save(); render();
  }
});

// ---- PIN lock: data is AES-GCM encrypted; key derived from PIN via PBKDF2 ----
function b64d(s){
  const bin = atob(s);
  const u = new Uint8Array(bin.length);
  for (let i = 0; i < bin.length; i++) u[i] = bin.charCodeAt(i);
  return u;
}
async function decryptWith(key){
  const pt = await crypto.subtle.decrypt({name:'AES-GCM', iv:b64d(ENC.iv)}, key, b64d(ENC.ct));
  return JSON.parse(new TextDecoder().decode(pt));
}
async function keyFromPin(pin){
  const mat = await crypto.subtle.importKey('raw', new TextEncoder().encode(pin), 'PBKDF2', false, ['deriveKey']);
  return crypto.subtle.deriveKey(
    {name:'PBKDF2', salt:b64d(ENC.salt), iterations:ENC.iter, hash:'SHA-256'},
    mat, {name:'AES-GCM', length:256}, true, ['decrypt']);
}
function setLocked(locked){
  document.getElementById('lock').hidden = !locked;
  document.querySelector('header').hidden = locked;
  document.getElementById('list').hidden = locked;
}
function startApp(){
  TOTAL = DATA.length;
  document.getElementById('total').textContent = TOTAL;
  setLocked(false);
  build(); load(); render();
}
let unlocking = false;
async function unlock(){
  if (unlocking) return;
  const el = document.getElementById('pin');
  const pin = el.value.replace(/[０-９]/g, ch => String.fromCharCode(ch.charCodeAt(0) - 0xFEE0)).trim();
  if (!pin) return;
  unlocking = true;
  const btn = document.getElementById('unlock');
  const err = document.getElementById('lockerr');
  err.hidden = true;
  btn.disabled = true;
  btn.textContent = '確認中…';
  try {
    if (!(window.crypto && crypto.subtle)) throw new Error('no-webcrypto');
    const key = await keyFromPin(pin);
    DATA = await decryptWith(key);
    const raw = new Uint8Array(await crypto.subtle.exportKey('raw', key));
    let bin = '';
    for (let i = 0; i < raw.length; i++) bin += String.fromCharCode(raw[i]);
    storeSet('co-check-key', btoa(bin));
    startApp();
  } catch(e){
    err.textContent = (e && e.message === 'no-webcrypto')
      ? 'このブラウザでは開けません。別のブラウザでお試しください'
      : 'PINが違います';
    err.hidden = false;
    el.value = '';
    el.focus();
  }
  btn.disabled = false;
  btn.textContent = '開く';
  unlocking = false;
}
document.getElementById('lockform').addEventListener('submit', e => { e.preventDefault(); unlock(); });
// touchend fires on the touchstart target even if the button shifts when the
// keyboard closes mid-tap; preventDefault suppresses the synthetic click.
document.getElementById('unlock').addEventListener('touchend', e => { e.preventDefault(); unlock(); });
// keyboard stays open: attempt automatically once the last digit is typed
document.getElementById('pin').addEventListener('input', e => {
  if (ENC.len && e.target.value.trim().length >= ENC.len) unlock();
});
// surface unexpected errors on the lock screen for remote diagnosis
window.addEventListener('error', ev => {
  const err = document.getElementById('lockerr');
  if (err && !document.getElementById('lock').hidden){
    err.textContent = 'エラー: ' + ev.message;
    err.hidden = false;
  }
});

(async () => {
  const cached = storeGet('co-check-key');
  if (cached){
    try {
      const key = await crypto.subtle.importKey('raw', b64d(cached), 'AES-GCM', false, ['decrypt']);
      DATA = await decryptWith(key);
      startApp();
      return;
    } catch(e){}
  }
  setLocked(true);
  document.getElementById('pin').focus();
})();
</script>
"""
html = html.replace('__ENC__', enc_js)
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'docs', 'index.html')
with open(out, 'w', encoding='utf-8') as f:
    f.write(html)
print('written', out, len(html))
